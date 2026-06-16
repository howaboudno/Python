# ==============================================================
# SCRIPT 1: setup_users.py
# Run this first to create the tournament and all 8 users.
# Run from your project root with your .venv active.
# ==============================================================

import sys
sys.path.insert(0, '.')

from backend.core.database import SessionLocal
from backend.core.security import hash_password
from backend.models.models import User, Tournament

PLAYERS = ["aboud", "ivan", "gosia", "antoine", "ridon", "petey", "alex", "elyas"]
PASSWORD = "WC2026!"

db = SessionLocal()

try:
    # Create tournament
    existing_tournament = db.query(Tournament).filter(Tournament.name == "World Cup 2026").first()
    if existing_tournament:
        print(f"Tournament already exists with id: {existing_tournament.id}")
        tournament_id = existing_tournament.id
    else:
        tournament = Tournament(
            name="World Cup 2026",
            tournament_type="World Cup",
            year=2026,
            is_active=True
        )
        db.add(tournament)
        db.commit()
        db.refresh(tournament)
        tournament_id = tournament.id
        print(f"Created tournament: World Cup 2026 (id: {tournament_id})")

    # Create users
    for username in PLAYERS:
        existing = db.query(User).filter(User.username == username).first()
        if existing:
            print(f"User already exists: {username}")
            continue
        user = User(
            username=username,
            hashed_password=hash_password(PASSWORD),
            is_admin=False
        )
        db.add(user)
        print(f"Created user: {username}")

    # Make aboud admin
    aboud = db.query(User).filter(User.username == "aboud").first()
    if aboud:
        aboud.is_admin = True
        print("Set aboud as admin")

    db.commit()
    print(f"\nDone. Tournament ID to use in import script: {tournament_id}")

except Exception as e:
    db.rollback()
    print(f"Error: {e}")
    raise
finally:
    db.close()


# ==============================================================
# SCRIPT 2: import_data.py
# Run this after setup_users.py.
# Reads the Excel file and imports:
#   - All 72 group stage fixtures (with kickoff times from worldcup26.ir)
#   - All predictions for all 8 players
#   - Group ranking predictions
#   - Bonus predictions (champion + top scorer)
# ==============================================================

"""
import sys
import requests
import pandas as pd
from datetime import datetime, timezone
sys.path.insert(0, '.')

from backend.core.database import SessionLocal
from backend.models.models import (
    User, Tournament, Fixture, FixturePrediction,
    GroupPrediction, BonusPredictions
)

EXCEL_PATH = "WK_2026_PREDICKS.xlsx"
TOURNAMENT_ID = 1  # Change if different
API_URL = "https://worldcup26.ir/get/games"

PLAYERS = ["aboud", "ivan", "gosia", "antoine", "ridon", "petey", "alex", "elyas"]

# Team name normalisation map — maps variants in the sheet to canonical names
TEAM_NAME_MAP = {
    "turkey": "Türkiye",
    "turkije": "Türkiye",
    "türkiye": "Türkiye",
    "usa": "United States",
    "united states": "United States",
    "bosnia & herz.": "Bosnia & Herz.",
    "bih": "Bosnia & Herz.",
    "biH": "Bosnia & Herz.",
    "bosnia": "Bosnia & Herz.",
    "nederland": "Netherlands",
    "nederland s": "Netherlands",
    "netherlands": "Netherlands",
    "ivory coast": "Ivory Coast",
    "cote divoire": "Ivory Coast",
    "cote d'ivoire": "Ivory Coast",
    "ivoire": "Ivory Coast",
    "south korea": "South Korea",
    "korea": "South Korea",
    "czech republic": "Czech Republic",
    "czechia": "Czech Republic",
    "czech": "Czech Republic",
    "dr congo": "DR Congo",
    "congo": "DR Congo",
    "saudi arabia": "Saudi Arabia",
    "saudi arabie": "Saudi Arabia",
    "saudi": "Saudi Arabia",
    "cape verde": "Cape Verde",
    "scotland": "Scotland",
    "schotland": "Scotland",
    "argentina": "Argentina",
    "austria": "Austria",
    "austra": "Austria",
    "algeria": "Algeria",
    "algeria ": "Algeria",
    "croatia ": "Croatia",
    "croatia": "Croatia",
    "ghana ": "Ghana",
    "ghana": "Ghana",
    "belgium": "Belgium",
    "belgie": "Belgium",
    "colombia": "Colombia",
    "uzbekistan": "Uzbekistan",
    "portugal": "Portugal",
    "england": "England",
    "france": "France",
    "spain": "Spain",
    "germany": "Germany",
    "brazil": "Brazil",
    "morocco": "Morocco",
    "mexico": "Mexico",
    "canada": "Canada",
    "japan": "Japan",
    "norway": "Norway",
    "ecuador": "Ecuador",
    "iran": "Iran",
    "iraq": "Iraq",
    "senegal": "Senegal",
    "uruguay": "Uruguay",
    "paraguay": "Paraguay",
    "australia": "Australia",
    "ghana": "Ghana",
    "panama": "Panama",
    "switzerland": "Switzerland",
    "qatar": "Qatar",
    "haiti": "Haiti",
    "new zealand": "New Zealand",
    "egypt": "Egypt",
    "jordan": "Jordan",
    "south africa": "South Africa",
    "sweden": "Sweden",
    "tunisia": "Tunisia",
    "curaçao": "Curaçao",
    "curacao": "Curaçao",
    "nigeria": "Nigeria",
    "ghana": "Ghana",
}

def normalise(name):
    if not name or str(name).strip() == "" or str(name) == "nan":
        return None
    return TEAM_NAME_MAP.get(str(name).strip().lower(), str(name).strip())

def fetch_fixtures_from_api():
    print("Fetching fixtures from worldcup26.ir...")
    resp = requests.get(API_URL, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    # Build a dict keyed by match number for quick lookup
    fixture_map = {}
    for match in data:
        match_num = match.get("matchNumber") or match.get("match_number") or match.get("id")
        kickoff = match.get("dateTime") or match.get("date") or match.get("kickoff")
        if match_num and kickoff:
            # Parse to datetime — handle both ISO and other formats
            try:
                dt = datetime.fromisoformat(str(kickoff).replace("Z", "+00:00"))
            except Exception:
                dt = datetime.strptime(str(kickoff)[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
            fixture_map[int(match_num)] = dt
    print(f"Got {len(fixture_map)} fixtures from API")
    return fixture_map

db = SessionLocal()

try:
    # Get all users
    users = {u.username: u for u in db.query(User).all()}
    missing = [p for p in PLAYERS if p not in users]
    if missing:
        print(f"Missing users: {missing}. Run setup_users.py first.")
        sys.exit(1)

    # Fetch kickoff times from API
    api_fixtures = fetch_fixtures_from_api()

    # Read Excel
    print("Reading Excel file...")
    df_matches = pd.read_excel(EXCEL_PATH, sheet_name="Group matches", header=0)
    df_pred = pd.read_excel(EXCEL_PATH, sheet_name="Predictions", header=None)

    # ── IMPORT FIXTURES ──────────────────────────────────────────
    print("Importing fixtures...")
    fixture_id_map = {}  # match_number -> db fixture id

    for _, row in df_matches.iterrows():
        match_num = int(row.iloc[1]) if pd.notna(row.iloc[1]) else None
        if not match_num:
            continue
        group = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
        team1 = normalise(row.iloc[4])
        team2 = normalise(row.iloc[6])

        kickoff = api_fixtures.get(match_num)
        if not kickoff:
            # Fall back to Excel date at noon UTC
            excel_date = row.iloc[3]
            if pd.notna(excel_date):
                kickoff = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(excel_date))
                kickoff = kickoff.to_pydatetime().replace(hour=12, tzinfo=timezone.utc)
            else:
                kickoff = datetime(2026, 6, 15, 12, 0, tzinfo=timezone.utc)

        existing = db.query(Fixture).filter(
            Fixture.tournament_id == TOURNAMENT_ID,
            Fixture.fixture_number == str(match_num)
        ).first()

        if existing:
            fixture_id_map[match_num] = existing.id
            continue

        fixture = Fixture(
            tournament_id=TOURNAMENT_ID,
            fixture_number=str(match_num),
            group=group,
            team_1=team1,
            team_2=team2,
            fixture_time=kickoff,
            stage="Group"
        )
        db.add(fixture)
        db.flush()
        fixture_id_map[match_num] = fixture.id

    db.commit()
    print(f"Imported {len(fixture_id_map)} fixtures")

    # ── PARSE PREDICTIONS FROM SHEET ─────────────────────────────
    # Player blocks start at col 11 (0-indexed: 11), each block is 9 cols wide
    # Within each block: +1=name shown in row 2, +3=pred_score1, +4=pred_score2
    # Row structure (0-indexed):
    #   Row 0: headers
    #   Row 1: player name row (col block+2 = name)
    #   Row 4+: match data rows

    # We'll re-read with openpyxl for more control
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Predictions"]

    # Player column offsets (0-indexed from col A=0)
    # Block starts: 11, 21, 31, 41, 51, 61, 71, 81
    BLOCK_START = [11, 21, 31, 41, 51, 61, 71, 81]
    # Within block: pred_team1=+1, pred_score1=+2, pred_score2=+3, pred_team2=+4
    # (based on sheet structure observed)

    # Find which row match data starts — scan for row where col B (idx 1) = 1
    # Match rows have M# in col B
    print("Parsing predictions from sheet...")

    rows = list(ws.iter_rows(values_only=True))

    # Find fixture rows by match number in col index 1
    match_rows = {}
    for i, row in enumerate(rows):
        val = row[1]
        if val and str(val).replace(".0","").isdigit():
            match_num = int(float(str(val)))
            if 1 <= match_num <= 72:
                match_rows[match_num] = i

    print(f"Found {len(match_rows)} match rows in sheet")

    # Import fixture predictions
    pred_count = 0
    for match_num, row_idx in match_rows.items():
        row = rows[row_idx]
        fixture_db_id = fixture_id_map.get(match_num)
        if not fixture_db_id:
            print(f"  No fixture in DB for match {match_num}, skipping")
            continue

        for p_idx, player in enumerate(PLAYERS):
            block = BLOCK_START[p_idx]
            s1 = row[block + 2]  # predicted score 1
            s2 = row[block + 3]  # predicted score 2

            if s1 is None or s2 is None:
                continue
            try:
                s1 = int(float(str(s1)))
                s2 = int(float(str(s2)))
            except (ValueError, TypeError):
                continue

            user_id = users[player].id

            existing = db.query(FixturePrediction).filter(
                FixturePrediction.user_id == user_id,
                FixturePrediction.fixture_id == fixture_db_id
            ).first()

            if existing:
                existing.predicted_score_1 = s1
                existing.predicted_score_2 = s2
            else:
                db.add(FixturePrediction(
                    user_id=user_id,
                    fixture_id=fixture_db_id,
                    predicted_score_1=s1,
                    predicted_score_2=s2,
                    predicted_pen_score_1=None,
                    predicted_pen_score_2=None
                ))
            pred_count += 1

    db.commit()
    print(f"Imported {pred_count} fixture predictions")

    # ── GROUP RANKING PREDICTIONS ─────────────────────────────────
    # These are in the QUALIFIED COUNTRIES section below the match rows
    # Find the row with "QUALIFIED COUNTRIES" header
    groups = list("ABCDEFGHIJKL")
    group_start_row = None
    for i, row in enumerate(rows):
        if any("QUALIFIED COUNTRIES" in str(c) for c in row if c):
            group_start_row = i
            break

    if group_start_row:
        print("Importing group predictions...")
        group_count = 0
        for g_idx, group in enumerate(groups):
            group_row = rows[group_start_row + 2 + g_idx]  # +2 for header rows

            for p_idx, player in enumerate(PLAYERS):
                block = BLOCK_START[p_idx]
                winner = normalise(group_row[block + 1])
                runner_up = normalise(group_row[block + 3])
                third = normalise(group_row[block + 5])

                if not winner and not runner_up and not third:
                    continue

                user_id = users[player].id
                existing = db.query(GroupPrediction).filter(
                    GroupPrediction.user_id == user_id,
                    GroupPrediction.tournament_id == TOURNAMENT_ID,
                    GroupPrediction.group == group
                ).first()

                if existing:
                    existing.first = winner or existing.first
                    existing.second = runner_up or existing.second
                    existing.third = third or existing.third
                else:
                    db.add(GroupPrediction(
                        user_id=user_id,
                        tournament_id=TOURNAMENT_ID,
                        group=group,
                        first=winner,
                        second=runner_up,
                        third=third
                    ))
                group_count += 1

        db.commit()
        print(f"Imported {group_count} group predictions")

    # ── BONUS PREDICTIONS ─────────────────────────────────────────
    print("Importing bonus predictions...")
    bonus_count = 0
    for i, row in enumerate(rows):
        if any("CHAMPION" in str(c) for c in row if c) and any("TOP SCORER" in str(rows[i+1][j]) for j in range(len(rows[i+1])) if rows[i+1][j]):
            champion_row = row
            top_scorer_row = rows[i + 1]

            for p_idx, player in enumerate(PLAYERS):
                block = BLOCK_START[p_idx]
                champion = normalise(champion_row[block + 1])
                top_scorer = normalise(top_scorer_row[block + 1])

                if not champion and not top_scorer:
                    continue

                user_id = users[player].id
                existing = db.query(BonusPredictions).filter(
                    BonusPredictions.user_id == user_id,
                    BonusPredictions.tournament_id == TOURNAMENT_ID
                ).first()

                if existing:
                    existing.predicted_winner = champion or existing.predicted_winner
                    existing.predicted_top_scorer = top_scorer or existing.predicted_top_scorer
                else:
                    db.add(BonusPredictions(
                        user_id=user_id,
                        tournament_id=TOURNAMENT_ID,
                        predicted_winner=champion,
                        predicted_top_scorer=top_scorer
                    ))
                bonus_count += 1
            break

    db.commit()
    print(f"Imported {bonus_count} bonus predictions")
    print("\\nAll done!")

except Exception as e:
    db.rollback()
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
    raise
finally:
    db.close()
"""
# The import script is wrapped in a docstring above so you can run setup_users.py
# independently first. To use import_data.py, save the content inside the
# triple quotes to a separate file called import_data.py
