import sys
import requests
import pandas as pd
from datetime import datetime, timezone
sys.path.insert(0, '.')

from backend.core.database import SessionLocal
from backend.models.models import (
    User, Tournament, Fixture, FixturePrediction,
    GroupPrediction, BonusPredictions
)

EXCEL_PATH = "WK_2026_PREDICKS.xlsx"
TOURNAMENT_ID = 1  # Change if setup_users.py printed a different ID
API_URL = "https://worldcup26.ir/get/games"

PLAYERS = ["aboud", "ivan", "gosia", "antoine", "ridon", "petey", "alex", "elyas"]

# Team name normalisation — maps sheet variants to canonical names
TEAM_NAME_MAP = {
    "turkey": "Türkiye",
    "turkije": "Türkiye",
    "türkiye": "Türkiye",
    "usa": "United States",
    "united states": "United States",
    "bosnia & herz.": "Bosnia & Herz.",
    "bih": "Bosnia & Herz.",
    "bosnia": "Bosnia & Herz.",
    "nederland": "Netherlands",
    "netherlands": "Netherlands",
    "ivory coast": "Ivory Coast",
    "cote divoire": "Ivory Coast",
    "cote d'ivoire": "Ivory Coast",
    "south korea": "South Korea",
    "korea": "South Korea",
    "czech republic": "Czech Republic",
    "czechia": "Czech Republic",
    "czech": "Czech Republic",
    "dr congo": "DR Congo",
    "congo": "DR Congo",
    "saudi arabia": "Saudi Arabia",
    "saudi arabie": "Saudi Arabia",
    "saudi": "Saudi Arabia",
    "cape verde": "Cape Verde",
    "schotland": "Scotland",
    "scotland": "Scotland",
    "curaçao": "Curaçao",
    "curacao": "Curaçao",
    "belgie": "Belgium",
    "belgium": "Belgium",
    "croatia ": "Croatia",
    "ghana ": "Ghana",
    "algeria ": "Algeria",
    "germany ": "Germany",
    "norway ": "Norway",
    "croatia": "Croatia",
    "ghana": "Ghana",
    "algeria": "Algeria",
}

def normalise(name):
    if not name or str(name).strip() in ("", "nan", "None"):
        return None
    cleaned = str(name).strip()
    return TEAM_NAME_MAP.get(cleaned.lower(), cleaned)

def fetch_fixtures_from_api():
    print("Fetching fixtures from worldcup26.ir...")
    try:
        resp = requests.get(API_URL, timeout=10)
        resp.raise_for_status()
        data = resp.json().get('games', [])
    except Exception as e:
        print(f"  API fetch failed: {e}. Will fall back to Excel dates.")
        return {}

    fixture_map = {}
    for match in data:
        match_num = match.get("id")
        kickoff = match.get("date") or match.get("dateTime") or match.get("kickoff")
        if match_num and kickoff:
            try:
                dt = datetime.fromisoformat(str(kickoff).replace("Z", "+00:00"))
            except Exception:
                try:
                    dt = datetime.strptime(str(kickoff)[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
                except Exception:
                    continue
            fixture_map[int(match_num)] = dt

    print(f"  Got {len(fixture_map)} kickoff times from API")
    return fixture_map
def excel_serial_to_datetime(serial):
    """Convert Excel date serial number to datetime."""
    try:
        base = pd.Timestamp("1899-12-30")
        dt = base + pd.Timedelta(days=int(float(serial)))
        return dt.to_pydatetime().replace(hour=18, tzinfo=timezone.utc)
    except Exception:
        return datetime(2026, 6, 15, 18, 0, tzinfo=timezone.utc)

db = SessionLocal()

try:
    # Verify users exist
    users = {u.username: u for u in db.query(User).all()}
    missing = [p for p in PLAYERS if p not in users]
    if missing:
        print(f"Missing users: {missing}. Run setup_users.py first.")
        sys.exit(1)
    print(f"Found all {len(PLAYERS)} users in database")

    # Fetch kickoff times
    api_fixtures = fetch_fixtures_from_api()

    # Read Excel sheets
    print("Reading Excel file...")
    df_matches = pd.read_excel(EXCEL_PATH, sheet_name="Group matches", header=0)

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Predictions"]
    rows = list(ws.iter_rows(values_only=True))

    # ── IMPORT FIXTURES ──────────────────────────────────────────
    print("\nImporting fixtures...")
    fixture_id_map = {}

    for _, row in df_matches.iterrows():
        match_num = row.iloc[1]
        if pd.isna(match_num):
            continue
        match_num = int(float(match_num))
        group = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
        team1 = normalise(row.iloc[4])
        team2 = normalise(row.iloc[7])  # was iloc[6]

        kickoff = api_fixtures.get(match_num)
        if not kickoff:
            date_val = row.iloc[3]
        if pd.notna(date_val):
            kickoff = pd.Timestamp(date_val).to_pydatetime().replace(hour=18, tzinfo=timezone.utc)
        else:
            kickoff = datetime(2026, 6, 15, 18, 0, tzinfo=timezone.utc)

        existing = db.query(Fixture).filter(
            Fixture.tournament_id == TOURNAMENT_ID,
            Fixture.fixture_number == str(match_num)
        ).first()

        if existing:
            fixture_id_map[match_num] = existing.id
            print(f"  Fixture {match_num} already exists, skipping")
            continue

        fixture = Fixture(
            tournament_id=TOURNAMENT_ID,
            fixture_number=str(match_num),
            group=group,
            team_1=team1,
            team_2=team2,
            fixture_time=kickoff,
            stage="Group"
        )
        db.add(fixture)
        db.flush()
        fixture_id_map[match_num] = fixture.id
        print(f"  Created fixture {match_num}: {team1} vs {team2} ({group})")

    db.commit()
    print(f"Fixtures done: {len(fixture_id_map)} total")

    # ── FIXTURE PREDICTIONS ───────────────────────────────────────
    # Player blocks in sheet (0-indexed columns):
    # Aboud=11, Ivan=21, Gosia=31, Antoine=41, Ridon=51, Petey=61, Alex=71, Elyas=81
    BLOCK_START = [11, 21, 31, 41, 51, 61, 71, 81]

    print("\nImporting fixture predictions...")

    # Find match rows by match number in column index 1
    match_rows = {}
    for i, row in enumerate(rows):
        val = row[1]
        if val is not None:
            try:
                match_num = int(float(str(val)))
                if 1 <= match_num <= 72:
                    match_rows[match_num] = i
            except (ValueError, TypeError):
                pass

    print(f"  Found {len(match_rows)} match rows")
    pred_count = 0

    for match_num, row_idx in sorted(match_rows.items()):
        row = rows[row_idx]
        fixture_db_id = fixture_id_map.get(match_num)
        if not fixture_db_id:
            print(f"  No DB fixture for match {match_num}, skipping predictions")
            continue

        for p_idx, player in enumerate(PLAYERS):
            block = BLOCK_START[p_idx]
            s1 = row[block + 2] if len(row) > block + 2 else None
            s2 = row[block + 3] if len(row) > block + 3 else None

            if s1 is None or s2 is None:
                continue
            try:
                s1 = int(float(str(s1)))
                s2 = int(float(str(s2)))
            except (ValueError, TypeError):
                continue

            user_id = users[player].id
            existing = db.query(FixturePrediction).filter(
                FixturePrediction.user_id == user_id,
                FixturePrediction.fixture_id == fixture_db_id
            ).first()

            if existing:
                existing.predicted_score_1 = s1
                existing.predicted_score_2 = s2
            else:
                db.add(FixturePrediction(
                    user_id=user_id,
                    fixture_id=fixture_db_id,
                    predicted_score_1=s1,
                    predicted_score_2=s2,
                    predicted_pen_score_1=None,
                    predicted_pen_score_2=None
                ))
            pred_count += 1

    db.commit()
    print(f"  Imported {pred_count} fixture predictions")

    # ── GROUP RANKING PREDICTIONS ─────────────────────────────────
    print("\nImporting group predictions...")
    groups = list("ABCDEFGHIJKL")

    # Find the QUALIFIED COUNTRIES section
    group_section_row = None
    for i, row in enumerate(rows):
        if any(c and "QUALIFIED COUNTRIES" in str(c) for c in row):
            group_section_row = i
            break

    if group_section_row is None:
        print("  Could not find QUALIFIED COUNTRIES section, skipping group predictions")
    else:
        # Header row after QUALIFIED COUNTRIES + 1 row for Winner/Runner-up labels
        # Groups start 2 rows below
        group_count = 0
        for g_idx, group in enumerate(groups):
            group_row_idx = group_section_row + 2 + g_idx
            if group_row_idx >= len(rows):
                break
            group_row = rows[group_row_idx]

            for p_idx, player in enumerate(PLAYERS):
                block = BLOCK_START[p_idx]
                # Within qualified countries block: +1=winner, +3=runner-up, +5=third
                winner = normalise(group_row[block + 1]) if len(group_row) > block + 1 else None
                runner_up = normalise(group_row[block + 3]) if len(group_row) > block + 3 else None
                third = normalise(group_row[block + 5]) if len(group_row) > block + 5 else None

                if not any([winner, runner_up, third]):
                    continue

                user_id = users[player].id
                existing = db.query(GroupPrediction).filter(
                    GroupPrediction.user_id == user_id,
                    GroupPrediction.tournament_id == TOURNAMENT_ID,
                    GroupPrediction.group_id == group
                ).first()

                if existing:
                    if winner: existing.first_place = winner
                    if runner_up: existing.second_place = runner_up
                    if third: existing.third_place = third
                else:
                    db.add(GroupPrediction(
                        user_id=user_id,
                        tournament_id=TOURNAMENT_ID,
                        group_id=group,
                        first_place=winner,
                        second_place=runner_up,
                        third_place=third
                    ))
                group_count += 1

        db.commit()
        print(f"  Imported {group_count} group predictions")

    # ── BONUS PREDICTIONS ─────────────────────────────────────────
    print("\nImporting bonus predictions...")
    bonus_count = 0

    for i, row in enumerate(rows):
        if any(c and "CHAMPION" in str(c) for c in row):
            champion_row = row
            top_scorer_row = rows[i + 1] if i + 1 < len(rows) else None

            if not top_scorer_row:
                break
            if not any(c and "TOP SCORER" in str(c) for c in top_scorer_row):
                continue

            for p_idx, player in enumerate(PLAYERS):
                block = BLOCK_START[p_idx]
                champion = normalise(champion_row[block + 1]) if len(champion_row) > block + 1 else None
                top_scorer = normalise(top_scorer_row[block + 1]) if len(top_scorer_row) > block + 1 else None

                if not champion and not top_scorer:
                    continue

                user_id = users[player].id
                existing = db.query(BonusPredictions).filter(
                    BonusPredictions.user_id == user_id,
                    BonusPredictions.tournament_id == TOURNAMENT_ID
                ).first()

                if existing:
                    if champion: existing.predicted_winner = champion
                    if top_scorer: existing.predicted_top_scorer = top_scorer
                else:
                    db.add(BonusPredictions(
                        user_id=user_id,
                        tournament_id=TOURNAMENT_ID,
                        predicted_winner=champion or "",
                        predicted_top_scorer=top_scorer or ""
                    ))
                bonus_count += 1
            break

    db.commit()
    print(f"  Imported {bonus_count} bonus predictions")

    print("\n✓ All data imported successfully!")
    print(f"  Tournament ID: {TOURNAMENT_ID}")
    print(f"  Users: {len(PLAYERS)}")
    print(f"  Fixtures: {len(fixture_id_map)}")
    print(f"  Predictions: {pred_count}")

except Exception as e:
    db.rollback()
    print(f"\nError: {e}")
    import traceback
    traceback.print_exc()
    raise
finally:
    db.close()
